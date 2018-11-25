import csv
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from module.fund import Fund, Transaction, T_Pool

''' SETTINGS '''
# GENERAL
today = datetime.today()
excel_file = 'Data.xlsx'
output_txt = 'output.txt'

# DEBUGGING
debug_mode = False

# USAGE TOOLS
output = open(output_txt, 'a')
transaction_list = []
excel_fund_list = []
fund_list = []

# EXCEL
wb = load_workbook(excel_file, data_only=True)  # Workbook
sheet_names = wb.sheetnames              # List of sheet names
ws_start = wb[sheet_names[0]]
ws_t = wb[sheet_names[2]]
ws_final = wb[sheet_names[3]]

''' FUNCTIONS '''
def debug(text):
    print('DEBUG: ' + str(text))

def output_me(text):
    output.write(text)
    output.write('\n')

''' MAIN '''
if debug_mode:
    # Used for debugging
    example_name = 'Fund A'

    fund_a = Fund(1, today - timedelta(days=40), example_name, 100, 1, 1, 100, 100)
    print('DEBUG: Initial units are ' + str(fund_a.u))

    if fund_a.f == example_name:
        print('1. Working')

    transaction_a = Transaction(1, [today, 'Fund A', 'Buy', 1000, 1, 1000])
    transaction_b = Transaction(1, [today + timedelta(days=40), 'Fund A1', 'Sell', 400, 1.1, 550])
    transaction_c = Transaction(1, [today + timedelta(days=41), 'Fund A1', 'Buy', 500, 1.2, 120])

    transaction_list = [transaction_a,transaction_b,transaction_c]

    for t in transaction_list:
        fund_a.transact(t)
        print(' ')

    print(fund_a.f)
    print(fund_a.u)
    print(fund_a.bc)

else:
    # Pull data from Excel into initial fund list
    for i, rows in enumerate(ws_start, start=0):
        if i == 0:
            # Miss the headers
            pass
        else:            
            if not rows[0].value:
                # Ignore blanks
                break
            current_fund = []
            for cell in rows:
                current_fund.append(cell.value)
            fund_list.append(Fund(current_fund[0], current_fund[1], current_fund[2],
                        current_fund[3], current_fund[4], current_fund[5],
                        current_fund[6], current_fund[7]))
            excel_fund_list.append(current_fund)
    
    # Cycle through transactions to apply to funds
    for i, rows in enumerate(ws_t, start=0):
        if i == 0:
            # Miss the headers
            pass
        else:
            if not rows[0].value:
                break
            current_t = []
            for cell in rows:
                current_t.append(cell.value)
            transaction_list.append(Transaction(current_t))
        
    for fund in fund_list:
        fund.debug()
    
    for tran in transaction_list:
        print(tran.k)

    for tran in transaction_list:
        new_fund = True
        for fund in fund_list:
            if new_fund == True:
                # Do if fund has not been found yet
                if fund.k == tran.k:
                    # Fund found so add transaction to it
                    new_fund = False
                    fund.transact(tran)
                else:
                    # Fund not already on list
                    pass
            else:
                # Fund not already on list
                pass
        if new_fund == True:
            # If fund was not found after cycling
            fund_list.append(Fund(tran.k, tran.d, tran.f, tran.u,
                            tran.p, None, None, tran.v))
        
    for fund in fund_list:
        fund.debug()